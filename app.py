from flask import Flask, render_template, request, redirect, session, flash
import sqlite3
from datetime import datetime
import os
import shutil
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import pagesizes
from flask import send_file
import io
from openpyxl import Workbook

app = Flask(__name__)
app.secret_key = "hospital_secret"
ADMIN_CLEAR_PASSWORD = "clear123"

DB = "hospital.db"

# ===================== DATABASE ======================

def query(q, args=(), one=False):
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(q, args)
    conn.commit()
    rv = cur.fetchall()
    conn.close()
    return (rv[0] if rv else None) if one else rv


def init_db():
    if os.path.exists(DB):
        return

    conn = sqlite3.connect(DB)
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE users(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT,
        password TEXT,
        role TEXT
    )
    """)

    departments = [
        "поліклініка","приймальне","педіатрія","інфекція",
        "терапія","неврологія","гінекологія","хірургія","ВАІТ"
    ]

    cur.execute("INSERT INTO users VALUES(NULL,'склад','1234','warehouse')")
    for d in departments:
        cur.execute("INSERT INTO users VALUES(NULL,?,?,?)",(d,d,"department"))

    cur.execute("""
    CREATE TABLE medicines(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE batches(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        medicine_id INTEGER,
        series TEXT,
        expiry TEXT,
        quantity INTEGER
    )
    """)

    cur.execute("""
    CREATE TABLE reservations(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        department TEXT,
        status TEXT,
        created TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE reservation_items(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        reservation_id INTEGER,
        medicine_id INTEGER,
        quantity INTEGER
    )
    """)

    cur.execute("""
    CREATE TABLE movement_log(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        medicine_id INTEGER,
        change INTEGER,
        type TEXT,
        created TEXT
    )
    """)
    cur.execute("""
    CREATE TABLE department_plan(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        medicine_id INTEGER,
        department TEXT,
        year INTEGER,
        plan INTEGER
    )
    """)
    conn.commit()
    conn.close()


init_db()
conn = sqlite3.connect(DB)
cur = conn.cursor()

try:
    cur.execute("ALTER TABLE medicines ADD COLUMN form TEXT")
except:
    pass

conn.commit()
conn.close()



# ===================== AUTH ======================

@app.route("/", methods=["GET","POST"])
def login():
    if request.method == "POST":
        user = query("SELECT * FROM users WHERE username=? AND password=?",
                     (request.form["username"], request.form["password"]),
                     one=True)
        if user:
            session["user"] = user["username"]
            session["role"] = user["role"]
            return redirect("/warehouse" if user["role"]=="warehouse" else "/department")
        else:
            flash("Невірний логін")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

# ===================== WAREHOUSE ======================

@app.route("/warehouse")
def warehouse():
    if session.get("role") != "warehouse":
        return redirect("/")

    medicines = query("""
    SELECT 
        m.id,
        m.name,
        IFNULL(SUM(b.quantity),0) as total,

        IFNULL((
            SELECT SUM(-change)
            FROM movement_log
            WHERE medicine_id = m.id
            AND type='out'
            AND created >= date('now','-60 day')
        ),0) as issued_2m

    FROM medicines m
    LEFT JOIN batches b ON m.id=b.medicine_id
    GROUP BY m.id
    """)

    batches = query("""
        SELECT b.*, m.name as medicine_name
        FROM batches b
        JOIN medicines m ON b.medicine_id=m.id
        ORDER BY expiry ASC
    """)

    reservations = query("""
        SELECT * FROM reservations
        ORDER BY id DESC
    """)

    return render_template(
        "warehouse.html",
        medicines=medicines,
        batches=batches,
        reservations=reservations,
        datetime=datetime
    )
    
@app.route("/add_batch", methods=["GET","POST"])
def add_batch():

    if session.get("role") != "warehouse":
        return redirect("/")

    if request.method == "POST":

        name = request.form["name"]
        form = request.form["form"]
        series = request.form["series"]
        expiry = request.form["expiry"]
        quantity = int(request.form["quantity"])

        med = query("SELECT * FROM medicines WHERE name=?", (name,), one=True)

        if med:
            med_id = med["id"]
        else:
            query("INSERT INTO medicines(name,form) VALUES(?,?)", (name,form))
            med_id = query("SELECT id FROM medicines WHERE name=?", (name,), one=True)["id"]

        query("""
        INSERT INTO batches(medicine_id,series,expiry,quantity)
        VALUES(?,?,?,?)
        """,(med_id,series,expiry,quantity))

        query("""
        INSERT INTO movement_log(medicine_id,change,type,created)
        VALUES(?,?,?,?)
        """,(med_id,quantity,"in",datetime.now()))

        return redirect("/warehouse")

    return render_template("add_batch.html")


@app.route("/import_excel", methods=["GET","POST"])
def import_excel():

    if session.get("role") != "warehouse":
        return redirect("/")

    if request.method == "POST":

        file = request.files["file"]

        from openpyxl import load_workbook

        wb = load_workbook(file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):

            name, form, series, expiry, quantity = row

            med = query("SELECT * FROM medicines WHERE name=?", (name,), one=True)

            if med:
                med_id = med["id"]
            else:
                query("INSERT INTO medicines(name,form) VALUES(?,?)",(name,form))
                med_id = query("SELECT id FROM medicines WHERE name=?", (name,), one=True)["id"]

            query("""
            INSERT INTO batches(medicine_id,series,expiry,quantity)
            VALUES(?,?,?,?)
            """,(med_id,series,expiry,quantity))

            query("""
            INSERT INTO movement_log(medicine_id,change,type,created)
            VALUES(?,?,?,?)
            """,(med_id,quantity,"in",datetime.now()))

        return redirect("/warehouse")

    return render_template("import_excel.html")



# ===================== DEPARTMENT ======================

@app.route("/department")
def department():

    if session.get("role") != "department":
        return redirect("/")

    department = session["user"]
    year = datetime.now().year

    med_list = query("""
        SELECT id, name, form
        FROM medicines
        ORDER BY name
    """)

    data = []

    for m in med_list:

        # ---------- склад ----------
        total = query("""
            SELECT IFNULL(SUM(quantity),0) as total
            FROM batches
            WHERE medicine_id=?
        """,(m["id"],),one=True)["total"]

        # ---------- план ----------
        plan = query("""
            SELECT plan
            FROM department_plan
            WHERE medicine_id=? AND department=? AND year=?
        """,(m["id"], department, year), one=True)

        plan_value = plan["plan"] if plan else 0

        # ---------- використано ----------
        used = query("""
            SELECT IFNULL(SUM(ri.quantity),0) as total
            FROM reservation_items ri
            JOIN reservations r ON ri.reservation_id=r.id
            WHERE ri.medicine_id=? 
            AND r.department=? 
            AND r.status='confirmed'
        """,(m["id"], department), one=True)["total"]

        # ---------- залишок плану ----------
        remaining = plan_value - used

        data.append({
            "id": m["id"],
            "name": m["name"],
            "form": m["form"],
            "total": total,
            "plan": plan_value,
            "used": used,
            "remaining": remaining
        })

    return render_template(
        "department.html",
        medicines=data
    )
@app.route("/create_reservation", methods=["POST"])
def create_reservation():

    if session.get("role") != "department":
        return redirect("/")

    department = session["user"]

    # створюємо накладну
    query("""
        INSERT INTO reservations(department,status,created)
        VALUES(?,?,?)
    """,(department,"pending",datetime.now()))

    reservation_id = query("""
        SELECT id FROM reservations
        ORDER BY id DESC
        LIMIT 1
    """,one=True)["id"]

    medicines = query("SELECT id FROM medicines")

    for m in medicines:

        qty = request.form.get(f"qty_{m['id']}")

        if qty and int(qty) > 0:

            query("""
                INSERT INTO reservation_items(reservation_id,medicine_id,quantity)
                VALUES(?,?,?)
            """,(reservation_id,m["id"],int(qty)))

    return redirect("/department")
# ===================== RESERVATIONS ======================

@app.route("/reservations")
def reservations():
    if session.get("role") != "warehouse":
        return redirect("/")

    data = query("SELECT * FROM reservations ORDER BY id DESC")
    return render_template("reservations.html", reservations=data)


@app.route("/reservation/<int:res_id>")
def reservation_detail(res_id):

    if session.get("role") != "warehouse":
        return redirect("/")

    reservation = query(
        "SELECT * FROM reservations WHERE id=?",
        (res_id,),
        one=True
    )

    items = query("""
    SELECT
        ri.id,
        ri.quantity,
        m.name,
        m.id AS medicine_id,

        IFNULL(
        (
            IFNULL(
                (
                    SELECT plan
                    FROM department_plan dp
                    WHERE dp.medicine_id = m.id
                    AND dp.department = (
                        SELECT department FROM reservations WHERE id=?
                    )
                    AND dp.year = strftime('%Y','now')
                ),0
            )
            -
            IFNULL(
                (
                    SELECT SUM(ri2.quantity)
                    FROM reservation_items ri2
                    JOIN reservations r2 ON ri2.reservation_id = r2.id
                    WHERE ri2.medicine_id = m.id
                    AND r2.department = (
                        SELECT department FROM reservations WHERE id=?
                    )
                    AND r2.status = 'confirmed'
                ),0
            )
        ),0) AS plan

    FROM reservation_items ri
    JOIN medicines m ON ri.medicine_id = m.id

    WHERE ri.reservation_id = ?
    """, (res_id, res_id, res_id))

    medicines = query("""
        SELECT m.id, m.name, IFNULL(SUM(b.quantity),0) as total
        FROM medicines m
        LEFT JOIN batches b ON m.id=b.medicine_id
        GROUP BY m.id
    """)

    return render_template(
        "reservation_detail.html",
        reservation=reservation,
        items=items,
        medicines=medicines
    )


@app.route("/update_reservation/<int:res_id>", methods=["POST"])
def update_reservation(res_id):

    if session.get("role") != "warehouse":
        return redirect("/")

    items = query(
        "SELECT * FROM reservation_items WHERE reservation_id=?",
        (res_id,)
    )

    for item in items:

        new_qty = request.form.get(f"qty_{item['id']}")

        if new_qty is None or int(new_qty) <= 0:

            query(
                "DELETE FROM reservation_items WHERE id=?",
                (item["id"],)
            )

        else:

            query(
                "UPDATE reservation_items SET quantity=? WHERE id=?",
                (int(new_qty), item["id"])
            )

    new_med = request.form.get("new_medicine")
    new_qty = request.form.get("new_quantity")

    if new_med and new_qty and int(new_qty) > 0:

        query(
            "INSERT INTO reservation_items VALUES(NULL,?,?,?)",
            (res_id, int(new_med), int(new_qty))
        )

    return redirect(f"/reservation/{res_id}")


@app.route("/confirm_reservation/<int:res_id>")
def confirm_reservation(res_id):

    if session.get("role") != "warehouse":
        return redirect("/")

    items = query(
        "SELECT * FROM reservation_items WHERE reservation_id=?",
        (res_id,)
    )

    for item in items:

        qty_needed = item["quantity"]

        total = query("""
            SELECT IFNULL(SUM(quantity),0) as total
            FROM batches
            WHERE medicine_id=?
        """, (item["medicine_id"],), one=True)["total"]

        if qty_needed > total:

            flash("Недостатньо препарату для списання")
            return redirect(f"/reservation/{res_id}")

        batches = query("""
            SELECT * FROM batches
            WHERE medicine_id=?
            ORDER BY expiry ASC
        """, (item["medicine_id"],))

        for b in batches:

            if qty_needed <= 0:
                break

            take = min(qty_needed, b["quantity"])
            new_qty = b["quantity"] - take

            query(
                "UPDATE batches SET quantity=? WHERE id=?",
                (new_qty, b["id"])
            )

            query(
                "INSERT INTO movement_log VALUES(NULL,?,?,?,?)",
                (item["medicine_id"], -take, "out", datetime.now())
            )

            qty_needed -= take

    query(
        "UPDATE reservations SET status='confirmed' WHERE id=?",
        (res_id,)
    )

    return redirect("/reservations")


@app.route("/reservation_pdf/<int:res_id>")
def reservation_pdf(res_id):

    if session.get("role") != "warehouse":
        return redirect("/")

    reservation = query(
        "SELECT * FROM reservations WHERE id=?",
        (res_id,),
        one=True
    )

    if not reservation or reservation["status"] != "confirmed":

        flash("PDF доступний тільки для підтверджених накладних")
        return redirect(f"/reservation/{res_id}")

    items = query("""
        SELECT m.name, ri.quantity
        FROM reservation_items ri
        JOIN medicines m ON ri.medicine_id = m.id
        WHERE ri.reservation_id=?
    """, (res_id,))

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagesizes.A4
    )

    elements = []

    styles = getSampleStyleSheet()

    elements.append(Paragraph("Hospital ERP - Накладна", styles["Title"]))
    elements.append(Spacer(1, 20))

    elements.append(Paragraph(
        f"Відділення: {reservation['department']}",
        styles["Normal"]
    ))

    elements.append(Paragraph(
        f"Дата: {reservation['created']}",
        styles["Normal"]
    ))

    elements.append(Spacer(1, 20))

    data = [["Назва препарату", "Кількість"]]

    for item in items:
        data.append([item["name"], str(item["quantity"])])

    table = Table(data, colWidths=[300, 100])

    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.grey),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("GRID", (0,0), (-1,-1), 1, colors.black),
        ("ALIGN",(1,1),(-1,-1),"CENTER")
    ]))

    elements.append(table)

    elements.append(Spacer(1, 30))

    elements.append(
        Paragraph("Підпис складу: ____________________", styles["Normal"])
    )

    doc.build(elements)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"reservation_{res_id}.pdf",
        mimetype="application/pdf"
    )


# ===================== MONTHLY REPORT ======================
# ===================== MONTHLY REPORT ======================

@app.route("/report_month")
def report_month():

    if session.get("role") != "warehouse":
        return redirect("/")

    medicines = query("SELECT * FROM medicines")

    wb = Workbook()
    ws = wb.active
    ws.title = "Місячний звіт"

    ws.append([
        "Препарат",
        "Залишок на початок місяця",
        "Прихід",
        "Видано відділенням",
        "Залишок на кінець"
    ])

    for m in medicines:

        start_balance = query("""
        SELECT IFNULL(SUM(quantity),0) as total
        FROM batches
        WHERE medicine_id=?
        """,(m["id"],),one=True)["total"]

        income = query("""
        SELECT IFNULL(SUM(change),0) as total
        FROM movement_log
        WHERE medicine_id=? AND type='in'
        """,(m["id"],),one=True)["total"]

        out = query("""
        SELECT IFNULL(SUM(change),0) as total
        FROM movement_log
        WHERE medicine_id=? AND type='out'
        """,(m["id"],),one=True)["total"]

        end_balance = start_balance + income + out

        ws.append([
            m["name"],
            start_balance,
            income,
            abs(out),
            end_balance
        ])

    file = io.BytesIO()
    wb.save(file)
    file.seek(0)

    return send_file(
        file,
        as_attachment=True,
        download_name="monthly_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ===================== WEEKLY REPORT ======================

@app.route("/report_week")
def report_week():

    if session.get("role") != "warehouse":
        return redirect("/")

    medicines = query("""
    SELECT m.name,
    IFNULL(SUM(b.quantity),0) as total
    FROM medicines m
    LEFT JOIN batches b ON m.id=b.medicine_id
    GROUP BY m.id
    """)

    wb = Workbook()
    ws = wb.active
    ws.title = "Тижневий звіт"

    ws.append(["Препарат","Поточний залишок"])

    for m in medicines:
        ws.append([m["name"], m["total"]])

    file = io.BytesIO()
    wb.save(file)
    file.seek(0)

    return send_file(
        file,
        as_attachment=True,
        download_name="weekly_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # ===================== EXCEL TEMPLATE ======================

@app.route("/excel_template")
def excel_template():

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    ws.append([
        "name",
        "form",
        "series",
        "expiry",
        "quantity"
    ])

    ws.append([
        "Парацетамол",
        "табл",
        "A123",
        "2027-05-01",
        "100"
    ])

    file = io.BytesIO()
    wb.save(file)
    file.seek(0)

    return send_file(
        file,
        as_attachment=True,
        download_name="excel_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# ===================== YEAR PLAN ======================

@app.route("/year_plan", methods=["GET","POST"])
def year_plan():

    if session.get("role") != "warehouse":
        return redirect("/")

    year = datetime.now().year

    # ---------- SAVE PLAN ----------
    if request.method == "POST":

        medicine_id = request.form["medicine_id"]
        department = request.form["department"]
        plan = request.form["plan"]

        existing = query("""
            SELECT id
            FROM department_plan
            WHERE medicine_id=? AND department=? AND year=?
        """, (medicine_id, department, year), one=True)

        if existing:
            query("""
                UPDATE department_plan
                SET plan=?
                WHERE id=?
            """, (plan, existing["id"]))
        else:
            query("""
                INSERT INTO department_plan(medicine_id, department, year, plan)
                VALUES(?,?,?,?)
            """, (medicine_id, department, year, plan))

        return redirect("/year_plan")

    # ---------- LOAD MEDICINES ----------
    medicines = query("""
        SELECT id, name, form
        FROM medicines
        ORDER BY name
    """)

    # ---------- LOAD DEPARTMENTS ----------
    departments = query("""
        SELECT username
        FROM users
        WHERE role='department'
        ORDER BY username
    """)

    data = []

    for m in medicines:

        row = {
            "id": m["id"],
            "name": m["name"],
            "form": m["form"],
            "departments": []
        }

        for d in departments:

            plan = query("""
                SELECT plan
                FROM department_plan
                WHERE medicine_id=? AND department=? AND year=?
            """, (m["id"], d["username"], year), one=True)

            plan_value = plan["plan"] if plan else 0

            used = query("""
                SELECT IFNULL(SUM(ri.quantity),0) as total
                FROM reservation_items ri
                JOIN reservations r ON ri.reservation_id=r.id
                WHERE ri.medicine_id=? AND r.department=? AND r.status='confirmed'
            """, (m["id"], d["username"]), one=True)["total"]

            remaining = plan_value - used

            row["departments"].append({
                "department": d["username"],
                "plan": plan_value,
                "remaining": remaining
            })

        data.append(row)

    return render_template(
        "year_plan.html",
        data=data,
        departments=departments,
        year=year
    )

    # ---------- BACKUP DATABASE ----------
    backup_folder = "backups"
    os.makedirs(backup_folder, exist_ok=True)

    backup_name = f"резервна_копія_аптека_{datetime.now().strftime('%Y_%m_%d_%H_%M')}.db"
    backup_path = os.path.join(backup_folder, backup_name)

    shutil.copy("hospital.db", backup_path)
    # -------------------------------------

    if password != ADMIN_CLEAR_PASSWORD:
        flash("Невірний пароль адміністратора")
        return redirect("/warehouse")

    query("DELETE FROM batches")
    query("DELETE FROM reservations")
    query("DELETE FROM reservation_items")
    query("DELETE FROM movement_log")

    flash(f"Склад очищено. Резервна копія створена: {backup_name}")
    return redirect("/warehouse")
    
    
@app.route("/create_backup", methods=["POST"])
def create_backup():

    if session.get("role") != "warehouse":
        return redirect("/")

    backup_folder = "backups"
    os.makedirs(backup_folder, exist_ok=True)

    backup_name = f"резервна_копія_аптека_{datetime.now().strftime('%Y_%m_%d_%H_%M')}.db"
    backup_path = os.path.join(backup_folder, backup_name)

    shutil.copy("hospital.db", backup_path)

    flash(f"Резервна копія створена: {backup_name}")
    return redirect("/warehouse")
# ===================== RUN ======================

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)